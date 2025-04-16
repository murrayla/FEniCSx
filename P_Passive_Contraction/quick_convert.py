import re
from openpyxl import Workbook

def extract_data_from_A(txt_path):
    with open(txt_path, 'r') as f:
        contents = f.read()

    array_matches = re.findall(r"\[([^\]]+)\]", contents)
    data = []
    for match in array_matches:
        floats = [float(val) for val in match.strip().split()]
        data.append(floats)
    return data

def extract_data_from_B(txt_path, group_size=5):
    with open(txt_path, 'r') as f:
        contents = f.read()

    # Extract floats
    floats = [float(val) for val in re.findall(r"[-+]?\d*\.\d+|\d+", contents)]

    rows = []
    current_row = []

    for i, val in enumerate(floats):
        if not current_row:
            current_row.append(val)
        elif val > current_row[-1]:
            current_row.append(val)
        else:
            # If not increasing, start a new row
            if len(current_row) >= group_size:
                rows.append(current_row[:group_size])
            current_row = [val]

        # If we reach group_size, commit row and start a new one
        if len(current_row) == group_size:
            rows.append(current_row)
            current_row = []

    # Add any remaining data if it's a valid increasing sequence
    if len(current_row) == group_size:
        rows.append(current_row)

    return rows

def write_to_excel(data_A, data_B, output_path, gap_columns=3):
    wb = Workbook()
    ws = wb.active

    # Write A data
    for row_idx, row in enumerate(data_A, start=1):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Write B data (after gap)
    start_col_B = len(data_A[0]) + gap_columns + 1
    for row_idx, row in enumerate(data_B, start=1):
        for col_idx, val in enumerate(row, start=start_col_B):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(output_path)

# === Main usage ===
data_A = extract_data_from_A('P_Passive_Contraction/c_1000_.txt')
data_B = extract_data_from_B('P_Passive_Contraction/m_1000_.txt')
write_to_excel(data_A, data_B, 'cm_1000.xlsx')